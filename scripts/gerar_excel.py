import json, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime

LOGO_PATH = os.path.join(os.path.dirname(__file__), 'sg_logo.png')

def gerar_excel(dados_json, output_path):
    d = json.loads(dados_json)
    empresa = d.get('empresa', 'SG Kussanguluca')
    periodo = d.get('periodo', '')
    total_receitas = float(d.get('totalReceitas', 0))
    total_despesas = float(d.get('totalDespesas', 0))
    receitas = d.get('receitas', [])
    despesas = d.get('despesas', [])

    wb = Workbook()

    # ─── CORES ────────────────────────────────────────────────
    VERDE_HEADER  = "1B5E20"
    VERDE_LIGHT   = "E8F5E9"
    VERMELHO_H    = "B71C1C"
    VERMELHO_L    = "FFEBEE"
    AZUL_HEADER   = "1A237E"
    AZUL_LIGHT    = "E8EAF6"
    CINZA_HEADER  = "37474F"
    CINZA_LIGHT   = "ECEFF1"
    BRANCO        = "FFFFFF"
    AMARELO_ALERTA= "FFF9C4"

    def borda(estilo="thin"):
        s = Side(style=estilo)
        return Border(left=s, right=s, top=s, bottom=s)

    def celula(ws, linha, col, valor, negrito=False, cor_fundo=None,
               cor_fonte=BRANCO, tamanho=11, alinhamento="left",
               formato=None, italico=False):
        c = ws.cell(row=linha, column=col, value=valor)
        c.font = Font(name="Arial", bold=negrito, size=tamanho,
                      color=cor_fonte, italic=italico)
        if cor_fundo:
            c.fill = PatternFill("solid", start_color=cor_fundo)
        c.alignment = Alignment(horizontal=alinhamento, vertical="center",
                                wrap_text=True)
        c.border = borda()
        if formato:
            c.number_format = formato
        return c

    def mesclar_titulo(ws, linha, c1, c2, texto, cor_fundo, cor_fonte=BRANCO, tamanho=12):
        ws.merge_cells(start_row=linha, start_column=c1,
                       end_row=linha, end_column=c2)
        c = ws.cell(row=linha, column=c1, value=texto)
        c.font = Font(name="Arial", bold=True, size=tamanho, color=cor_fonte)
        c.fill = PatternFill("solid", start_color=cor_fundo)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda()
        ws.row_dimensions[linha].height = 22

    fmt_kz = '#,##0.00" Kz"'

    # ══════════════════════════════════════════════════════════
    # ABA 1 — RESUMO
    # ══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumo"
    ws1.sheet_view.showGridLines = False

    # Cabeçalho principal
    ws1.merge_cells("A1:F1")
    c = ws1["A1"]
    c.value = "SG KUSSANGULUCA — RELATÓRIO FINANCEIRO"
    c.font = Font(name="Arial", bold=True, size=16, color=BRANCO)
    c.fill = PatternFill("solid", start_color=AZUL_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells("A2:F2")
    c = ws1["A2"]
    c.value = f"Empresa: {empresa}   |   Período: {periodo}   |   Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font = Font(name="Arial", size=10, color="000000", italic=True)
    c.fill = PatternFill("solid", start_color=AZUL_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 18

    ws1.row_dimensions[3].height = 10  # espaço

    # Logo no canto superior direito
    if os.path.exists(LOGO_PATH):
        logo = XLImage(LOGO_PATH)
        logo.width  = 52
        logo.height = 52
        logo.anchor = 'F1'
        ws1.add_image(logo)
        ws2 = None  # placeholder, real ws2 created later

    # Indicadores
    mesclar_titulo(ws1, 4, 1, 6, "📊  INDICADORES FINANCEIROS", CINZA_HEADER, tamanho=12)

    headers_ind = ["Indicador", "Valor", "", "Indicador", "Valor", ""]
    for i, h in enumerate(headers_ind, 1):
        if h:
            celula(ws1, 5, i, h, negrito=True, cor_fundo=CINZA_HEADER, tamanho=10, alinhamento="center")

    saldo = total_receitas - total_despesas
    margem = (saldo / total_receitas * 100) if total_receitas else 0
    razao  = (total_despesas / total_receitas * 100) if total_receitas else 0
    n_rec  = len(receitas)
    n_desp = len(despesas)

    indicadores = [
        ("Total de Receitas", total_receitas, "Total de Despesas", total_despesas),
        ("Saldo Líquido",      saldo,          "Margem de Lucro",  margem / 100),
        ("Razão Despesa/Rec.", razao / 100,    "Nº de Receitas",   n_rec),
        ("Nº de Despesas",     n_desp,         "Nº Total Transações", n_rec + n_desp),
    ]

    fmts = [fmt_kz, fmt_kz, fmt_kz, "0.0%", "0.0%", "0", "0", "0"]
    row = 6
    for i, (l1, v1, l2, v2) in enumerate(indicadores):
        cor_v1 = VERDE_LIGHT if i == 0 else (VERMELHO_L if i == 1 else AZUL_LIGHT)
        cor_v2 = VERMELHO_L if i == 0 else AZUL_LIGHT
        celula(ws1, row, 1, l1, negrito=True, cor_fundo=CINZA_LIGHT, cor_fonte="000000", tamanho=10)
        celula(ws1, row, 2, v1, cor_fundo=cor_v1, cor_fonte="000000", tamanho=10,
               alinhamento="right", formato=fmts[i*2])
        ws1.cell(row=row, column=3).border = borda()
        celula(ws1, row, 4, l2, negrito=True, cor_fundo=CINZA_LIGHT, cor_fonte="000000", tamanho=10)
        celula(ws1, row, 5, v2, cor_fundo=cor_v2, cor_fonte="000000", tamanho=10,
               alinhamento="right", formato=fmts[i*2+1])
        ws1.cell(row=row, column=6).border = borda()
        ws1.row_dimensions[row].height = 20
        row += 1

    row += 1  # espaço

    # Alerta de saúde financeira
    if razao > 80:
        mesclar_titulo(ws1, row, 1, 6,
            f"⚠️  ALERTA: Despesas representam {razao:.1f}% das receitas. Recomenda-se reduzir gastos.",
            "FF8F00", cor_fonte=BRANCO, tamanho=10)
        ws1.row_dimensions[row].height = 20
        row += 1
    elif saldo > 0:
        mesclar_titulo(ws1, row, 1, 6,
            f"✅  Situação financeira POSITIVA. Saldo favorável de {saldo:,.2f} Kz.",
            VERDE_HEADER, cor_fonte=BRANCO, tamanho=10)
        ws1.row_dimensions[row].height = 20
        row += 1

    # Larguras
    for col, w in zip(range(1, 7), [28, 18, 4, 28, 18, 4]):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ══════════════════════════════════════════════════════════
    # ABA 2 — RECEITAS
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Receitas")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:E1")
    c = ws2["A1"]
    c.value = f"SG KUSSANGULUCA — RECEITAS DETALHADAS   |   {empresa}   |   {periodo}"
    c.font = Font(name="Arial", bold=True, size=13, color=BRANCO)
    c.fill = PatternFill("solid", start_color=VERDE_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    cabecalhos_rec = ["#", "Data", "Descrição", "Categoria", "Valor (Kz)"]
    for i, h in enumerate(cabecalhos_rec, 1):
        celula(ws2, 2, i, h, negrito=True, cor_fundo=VERDE_HEADER, tamanho=10, alinhamento="center")

    for idx, r in enumerate(receitas, 1):
        cor = VERDE_LIGHT if idx % 2 == 0 else BRANCO
        celula(ws2, idx+2, 1, idx,    cor_fundo=cor, cor_fonte="000000", alinhamento="center")
        celula(ws2, idx+2, 2, r.get('data',''), cor_fundo=cor, cor_fonte="000000")
        celula(ws2, idx+2, 3, r.get('descricao',''), cor_fundo=cor, cor_fonte="000000")
        celula(ws2, idx+2, 4, r.get('categoria',''), cor_fundo=cor, cor_fonte="000000")
        celula(ws2, idx+2, 5, float(r.get('valor',0)), cor_fundo=cor, cor_fonte="000000",
               alinhamento="right", formato=fmt_kz)

    total_row = len(receitas) + 3
    ws2.merge_cells(f"A{total_row}:D{total_row}")
    celula(ws2, total_row, 1, "TOTAL RECEITAS", negrito=True, cor_fundo=VERDE_HEADER, tamanho=11, alinhamento="right")
    celula(ws2, total_row, 5, f"=SUM(E3:E{total_row-1})", negrito=True,
           cor_fundo=VERDE_HEADER, tamanho=11, alinhamento="right", formato=fmt_kz)
    ws2.row_dimensions[total_row].height = 22

    for col, w in zip(range(1,6), [5, 14, 35, 20, 18]):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ══════════════════════════════════════════════════════════
    # ABA 3 — DESPESAS
    # ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Despesas")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:E1")
    c = ws3["A1"]
    c.value = f"SG KUSSANGULUCA — DESPESAS DETALHADAS   |   {empresa}   |   {periodo}"
    c.font = Font(name="Arial", bold=True, size=13, color=BRANCO)
    c.fill = PatternFill("solid", start_color=VERMELHO_H)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 32

    cabecalhos_desp = ["#", "Data", "Descrição", "Categoria", "Valor (Kz)"]
    for i, h in enumerate(cabecalhos_desp, 1):
        celula(ws3, 2, i, h, negrito=True, cor_fundo=VERMELHO_H, tamanho=10, alinhamento="center")

    for idx, d2 in enumerate(despesas, 1):
        cor = VERMELHO_L if idx % 2 == 0 else BRANCO
        celula(ws3, idx+2, 1, idx,      cor_fundo=cor, cor_fonte="000000", alinhamento="center")
        celula(ws3, idx+2, 2, d2.get('data',''), cor_fundo=cor, cor_fonte="000000")
        celula(ws3, idx+2, 3, d2.get('descricao',''), cor_fundo=cor, cor_fonte="000000")
        celula(ws3, idx+2, 4, d2.get('categoria',''), cor_fundo=cor, cor_fonte="000000")
        celula(ws3, idx+2, 5, float(d2.get('valor',0)), cor_fundo=cor, cor_fonte="000000",
               alinhamento="right", formato=fmt_kz)

    total_row3 = len(despesas) + 3
    ws3.merge_cells(f"A{total_row3}:D{total_row3}")
    celula(ws3, total_row3, 1, "TOTAL DESPESAS", negrito=True, cor_fundo=VERMELHO_H, tamanho=11, alinhamento="right")
    celula(ws3, total_row3, 5, f"=SUM(E3:E{total_row3-1})", negrito=True,
           cor_fundo=VERMELHO_H, tamanho=11, alinhamento="right", formato=fmt_kz)
    ws3.row_dimensions[total_row3].height = 22

    for col, w in zip(range(1,6), [5, 14, 35, 20, 18]):
        ws3.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)
    print(f"Excel gerado: {output_path}")

if __name__ == "__main__":
    # sys.argv[1] = ficheiro JSON de input, sys.argv[2] = ficheiro de output
    import json
    with open(sys.argv[1], 'r', encoding='utf-8') as f:
        dados_json = f.read()
    gerar_excel(dados_json, sys.argv[2])